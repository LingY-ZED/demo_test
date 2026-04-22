"""
数据导入服务
支持 Excel/CSV 文件导入
"""

import csv
from datetime import datetime
from decimal import Decimal
from typing import Dict, List, Any
from pathlib import Path

from openpyxl import load_workbook


class TableFormatError(ValueError):
    """上传表格格式错误"""


class UploadService:
    """数据导入服务"""

    # CSV 编码
    ENCODING = "utf-8-sig"

    @classmethod
    def parse_transactions(cls, file_path: str, case_id: int) -> List[Dict[str, Any]]:
        """
        解析资金流水文件

        支持格式：
        - 交易发生时间,打款方 (账号/姓名),收款方 (账号/姓名),交易金额 (元),支付方式,交易备注 / 转账留言
        """
        records = []
        suffix = Path(file_path).suffix.lower()

        if suffix in [".xlsx", ".xls"]:
            records = cls._parse_excel_transactions(file_path, case_id)
        else:
            records = cls._parse_csv_transactions(file_path, case_id)

        cls._ensure_non_empty_records(records, "资金流水")

        return records

    @classmethod
    def _parse_csv_transactions(
        cls, file_path: str, case_id: int
    ) -> List[Dict[str, Any]]:
        """解析CSV资金流水"""
        records = []
        with open(file_path, encoding=cls.ENCODING) as f:
            reader = csv.DictReader(f)
            cls._validate_required_columns(
                reader.fieldnames,
                [
                    ["交易发生时间"],
                    ["打款方", "打款方 (账号/姓名)"],
                    ["收款方", "收款方 (账号/姓名)"],
                    ["交易金额", "交易金额 (元)"],
                ],
                "资金流水",
            )
            for row in reader:
                try:
                    record = {
                        "case_id": case_id,
                        "transaction_time": cls._parse_datetime(
                            row.get("交易发生时间", "")
                        ),
                        "payer": row.get("打款方 (账号/姓名)", "").strip(),
                        "payee": row.get("收款方 (账号/姓名)", "").strip(),
                        "amount": Decimal(
                            row.get("交易金额 (元)", "0").replace(",", "")
                        ),
                        "payment_method": row.get("支付方式", "").strip() or None,
                        "remark": row.get("交易备注 / 转账留言", "").strip() or None,
                    }
                    records.append(record)
                except Exception as e:
                    print(f"解析资金流水行失败: {e}, 数据: {row}")
        return records

    @classmethod
    def _parse_excel_transactions(
        cls, file_path: str, case_id: int
    ) -> List[Dict[str, Any]]:
        """解析Excel资金流水"""
        records = []
        wb = load_workbook(file_path, data_only=True)
        ws = wb.active

        # 获取表头
        headers = [cell.value for cell in ws[1]]
        cls._validate_required_columns(
            headers,
            [
                ["交易发生时间"],
                ["打款方", "打款方 (账号/姓名)"],
                ["收款方", "收款方 (账号/姓名)"],
                ["交易金额", "交易金额 (元)"],
            ],
            "资金流水",
        )

        for row in ws.iter_rows(min_row=2, values_only=True):
            try:
                data = dict(zip(headers, row))
                record = {
                    "case_id": case_id,
                    "transaction_time": cls._parse_datetime(
                        data.get("交易发生时间", "")
                    ),
                    "payer": str(
                        data.get("打款方") or data.get("打款方 (账号/姓名)", "")
                    ).strip(),
                    "payee": str(
                        data.get("收款方") or data.get("收款方 (账号/姓名)", "")
                    ).strip(),
                    "amount": Decimal(
                        str(
                            data.get("交易金额") or data.get("交易金额 (元)", "0")
                        ).replace(",", "")
                    ),
                    "payment_method": str(data.get("支付方式", "")).strip() or None,
                    "remark": str(
                        data.get("交易备注") or data.get("交易备注 / 转账留言", "")
                    ).strip()
                    or None,
                }
                records.append(record)
            except Exception as e:
                print(f"解析Excel资金流水行失败: {e}")
        return records

    @classmethod
    def parse_communications(cls, file_path: str, case_id: int) -> List[Dict[str, Any]]:
        """解析通讯记录文件"""
        records = []
        suffix = Path(file_path).suffix.lower()

        if suffix in [".xlsx", ".xls"]:
            records = cls._parse_excel_communications(file_path, case_id)
        else:
            records = cls._parse_csv_communications(file_path, case_id)

        cls._ensure_non_empty_records(records, "通讯记录")

        return records

    @classmethod
    def _parse_csv_communications(
        cls, file_path: str, case_id: int
    ) -> List[Dict[str, Any]]:
        """解析CSV通讯记录"""
        records = []
        with open(file_path, encoding=cls.ENCODING) as f:
            reader = csv.DictReader(f)
            cls._validate_required_columns(
                reader.fieldnames,
                [
                    ["联络时间"],
                    ["发起方 (微信号/姓名)"],
                    ["接收方 (微信号/姓名)"],
                    ["聊天内容"],
                ],
                "通讯记录",
            )
            for row in reader:
                try:
                    record = {
                        "case_id": case_id,
                        "communication_time": cls._parse_datetime(
                            row.get("联络时间", "")
                        ),
                        "initiator": row.get("发起方 (微信号/姓名)", "").strip(),
                        "receiver": row.get("接收方 (微信号/姓名)", "").strip(),
                        "content": row.get("聊天内容", "").strip() or None,
                    }
                    records.append(record)
                except Exception as e:
                    print(f"解析通讯记录行失败: {e}")
        return records

    @classmethod
    def _parse_excel_communications(
        cls, file_path: str, case_id: int
    ) -> List[Dict[str, Any]]:
        """解析Excel通讯记录"""
        records = []
        wb = load_workbook(file_path, data_only=True)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        cls._validate_required_columns(
            headers,
            [
                ["联络时间"],
                ["发起方 (微信号/姓名)"],
                ["接收方 (微信号/姓名)"],
                ["聊天内容"],
            ],
            "通讯记录",
        )

        for row in ws.iter_rows(min_row=2, values_only=True):
            try:
                data = dict(zip(headers, row))
                record = {
                    "case_id": case_id,
                    "communication_time": cls._parse_datetime(data.get("联络时间", "")),
                    "initiator": str(data.get("发起方 (微信号/姓名)", "")).strip(),
                    "receiver": str(data.get("接收方 (微信号/姓名)", "")).strip(),
                    "content": str(data.get("聊天内容", "")).strip() or None,
                }
                records.append(record)
            except Exception as e:
                print(f"解析Excel通讯记录行失败: {e}")
        return records

    @classmethod
    def parse_logistics(cls, file_path: str, case_id: int) -> List[Dict[str, Any]]:
        """解析物流记录文件"""
        records = []
        suffix = Path(file_path).suffix.lower()

        if suffix in [".xlsx", ".xls"]:
            records = cls._parse_excel_logistics(file_path, case_id)
        else:
            records = cls._parse_csv_logistics(file_path, case_id)

        cls._ensure_non_empty_records(records, "物流记录")

        return records

    @classmethod
    def _parse_csv_logistics(cls, file_path: str, case_id: int) -> List[Dict[str, Any]]:
        """解析CSV物流记录"""
        records = []
        with open(file_path, encoding=cls.ENCODING) as f:
            reader = csv.DictReader(f)
            cls._validate_required_columns(
                reader.fieldnames,
                [
                    ["发货时间"],
                    ["发件人/网点"],
                    ["收件人/地址"],
                ],
                "物流记录",
            )
            for row in reader:
                try:
                    # 解析发件人/收件人（可能包含地址）
                    sender = row.get("发件人/网点", "").strip()
                    receiver = row.get("收件人/地址", "").strip()

                    record = {
                        "case_id": case_id,
                        "shipping_time": cls._parse_datetime(row.get("发货时间", "")),
                        "tracking_no": row.get("快递单号", "").strip() or None,
                        "sender": cls._extract_name(sender),
                        "sender_address": cls._extract_address(sender),
                        "receiver": cls._extract_name(receiver),
                        "receiver_address": cls._extract_address(receiver),
                        "description": row.get("寄件物品描述", "").strip() or None,
                        "weight": cls._parse_decimal(row.get("包裹重量(公斤)", "")),
                    }
                    records.append(record)
                except Exception as e:
                    print(f"解析物流记录行失败: {e}")
        return records

    @classmethod
    def _parse_excel_logistics(
        cls, file_path: str, case_id: int
    ) -> List[Dict[str, Any]]:
        """解析Excel物流记录"""
        records = []
        wb = load_workbook(file_path, data_only=True)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        cls._validate_required_columns(
            headers,
            [
                ["发货时间"],
                ["发件人/网点"],
                ["收件人/地址"],
            ],
            "物流记录",
        )

        for row in ws.iter_rows(min_row=2, values_only=True):
            try:
                data = dict(zip(headers, row))
                sender = str(data.get("发件人/网点", "")).strip()
                receiver = str(data.get("收件人/地址", "")).strip()

                record = {
                    "case_id": case_id,
                    "shipping_time": cls._parse_datetime(data.get("发货时间", "")),
                    "tracking_no": str(data.get("快递单号", "")).strip() or None,
                    "sender": cls._extract_name(sender),
                    "sender_address": cls._extract_address(sender),
                    "receiver": cls._extract_name(receiver),
                    "receiver_address": cls._extract_address(receiver),
                    "description": str(data.get("寄件物品描述", "")).strip() or None,
                    "weight": cls._parse_decimal(data.get("包裹重量(公斤)", "")),
                }
                records.append(record)
            except Exception as e:
                print(f"解析Excel物流记录行失败: {e}")
        return records

    @staticmethod
    def _parse_datetime(value: str) -> datetime:
        """解析日期时间"""
        if not value:
            return datetime.now()
        formats = [
            "%Y-%m-%d %H:%M",
            "%Y-%m-%d",
            "%Y/%m/%d %H:%M",
            "%Y/%m/%d",
        ]
        for fmt in formats:
            try:
                return datetime.strptime(str(value).strip(), fmt)
            except ValueError:
                continue
        return datetime.now()

    @staticmethod
    def _parse_decimal(value: str) -> Decimal:
        """解析Decimal"""
        if not value:
            return Decimal("0")
        try:
            return Decimal(str(value).replace(",", "").strip())
        except Exception:
            return Decimal("0")

    @staticmethod
    def _extract_name(text: str) -> str:
        """从'姓名 (地址)'格式中提取姓名"""
        if "(" in text:
            return text.split("(")[0].strip()
        return text.strip()

    @staticmethod
    def _extract_address(text: str) -> str:
        """从'姓名 (地址)'格式中提取地址"""
        if "(" in text and ")" in text:
            start = text.index("(")
            end = text.index(")")
            return text[start + 1 : end].strip()
        return ""

    @staticmethod
    def _normalize_headers(headers: List[Any]) -> List[str]:
        """归一化表头，便于模板列名校验"""
        if not headers:
            return []
        result = []
        for header in headers:
            if header is None:
                continue
            text = str(header).strip()
            if text:
                result.append(text)
        return result

    @classmethod
    def _validate_required_columns(
        cls, headers: List[Any], required_groups: List[List[str]], table_name: str
    ) -> None:
        """校验上传模板的必填列是否存在（支持同义列名）"""
        normalized_headers = set(cls._normalize_headers(headers))
        if not normalized_headers:
            raise TableFormatError(f"{table_name}表格格式错误：缺少表头")

        missing_columns = []
        for group in required_groups:
            if not any(candidate in normalized_headers for candidate in group):
                missing_columns.append("/".join(group))

        if missing_columns:
            raise TableFormatError(
                f"{table_name}表格格式错误：缺少必填列 {', '.join(missing_columns)}"
            )

    @staticmethod
    def _ensure_non_empty_records(
        records: List[Dict[str, Any]], table_name: str
    ) -> None:
        """上传成功但未读取到有效行时，提示用户检查模板与数据内容"""
        if not records:
            raise TableFormatError(
                f"{table_name}表格格式错误：未读取到有效数据，请检查列名和内容后重试"
            )
