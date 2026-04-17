"""
数据导出工具
支持CSV/Excel导出，自定义筛选条件
"""
import csv
import io
import os
from typing import Optional, List, Dict, Any
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


class Exporter:
    """数据导出工具类"""

    # 默认导出目录
    DEFAULT_EXPORT_DIR = os.path.join(os.path.dirname(__file__), '..', 'data', 'exports')

    @classmethod
    def export_to_csv(
        cls,
        data: List[Dict[str, Any]],
        headers: List[str],
        filename: Optional[str] = None,
        output_dir: Optional[str] = None
    ) -> str:
        """
        导出数据到CSV文件

        Args:
            data: 数据列表
            headers: 表头列表
            filename: 文件名（不含扩展名）
            output_dir: 输出目录

        Returns:
            导出文件路径
        """
        if filename is None:
            filename = f"export_{datetime.now().strftime('%Y%m%d%H%M%S')}"
        if not filename.endswith('.csv'):
            filename += '.csv'

        if output_dir is None:
            output_dir = cls.DEFAULT_EXPORT_DIR
        os.makedirs(output_dir, exist_ok=True)

        filepath = os.path.join(output_dir, filename)

        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(headers)

        for row in data:
            row_data = []
            for header in headers:
                value = row.get(header, '')
                if isinstance(value, (list, dict)):
                    value = str(value)
                row_data.append(value)
            writer.writerow(row_data)

        with open(filepath, 'w', encoding='utf-8-sig', newline='') as f:
            f.write(output.getvalue())

        return filepath

    @classmethod
    def export_to_excel(
        cls,
        data: List[Dict[str, Any]],
        headers: List[str],
        filename: Optional[str] = None,
        output_dir: Optional[str] = None,
        sheet_name: str = "数据导出"
    ) -> str:
        """
        导出数据到Excel文件

        Args:
            data: 数据列表
            headers: 表头列表
            filename: 文件名（不含扩展名）
            output_dir: 输出目录
            sheet_name: 工作表名称

        Returns:
            导出文件路径
        """
        if not HAS_OPENPYXL:
            # 降级为CSV
            return cls.export_to_csv(data, headers, filename, output_dir)

        if filename is None:
            filename = f"export_{datetime.now().strftime('%Y%m%d%H%M%S')}"
        if not filename.endswith('.xlsx'):
            filename += '.xlsx'

        if output_dir is None:
            output_dir = cls.DEFAULT_EXPORT_DIR
        os.makedirs(output_dir, exist_ok=True)

        filepath = os.path.join(output_dir, filename)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name

        # 写入表头
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

        # 写入数据
        for row_idx, row_data in enumerate(data, 2):
            for col_idx, header in enumerate(headers, 1):
                value = row_data.get(header, '')
                if isinstance(value, (list, dict)):
                    value = str(value)
                ws.cell(row=row_idx, column=col_idx, value=value)

        # 自动调整列宽
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width

        wb.save(filepath)
        return filepath

    @classmethod
    def export_transactions(
        cls,
        transactions: List[Dict[str, Any]],
        format: str = 'csv',
        filename: Optional[str] = None
    ) -> str:
        """
        导出资金流水数据

        Args:
            transactions: 资金流水列表
            format: 导出格式 ('csv' 或 'excel')
            filename: 文件名

        Returns:
            导出文件路径
        """
        headers = ["ID", "案件ID", "案件编号", "交易时间", "打款方", "收款方", "金额", "支付方式", "备注"]
        data = []
        for t in transactions:
            data.append({
                "ID": t.get('id', ''),
                "案件ID": t.get('case_id', ''),
                "案件编号": t.get('case_no', ''),
                "交易时间": t.get('transaction_time', ''),
                "打款方": t.get('payer', ''),
                "收款方": t.get('payee', ''),
                "金额": t.get('amount', ''),
                "支付方式": t.get('payment_method', ''),
                "备注": t.get('remark', '')
            })

        if format == 'excel':
            return cls.export_to_excel(data, headers, filename)
        return cls.export_to_csv(data, headers, filename)

    @classmethod
    def export_persons(
        cls,
        persons: List[Dict[str, Any]],
        format: str = 'csv',
        filename: Optional[str] = None
    ) -> str:
        """
        导出人员台账数据

        Args:
            persons: 人员列表
            format: 导出格式
            filename: 文件名

        Returns:
            导出文件路径
        """
        headers = ["ID", "姓名", "角色", "是否有授权", "主观明知评分", "非法经营数额", "关联案件数"]
        data = []
        for p in persons:
            data.append({
                "ID": p.get('id', ''),
                "姓名": p.get('name', ''),
                "角色": p.get('role', ''),
                "是否有授权": "是" if p.get('is_authorized') else "否",
                "主观明知评分": p.get('subjective_knowledge_score', ''),
                "非法经营数额": p.get('illegal_business_amount', ''),
                "关联案件数": p.get('linked_cases', '')
            })

        if format == 'excel':
            return cls.export_to_excel(data, headers, filename)
        return cls.export_to_csv(data, headers, filename)

    @classmethod
    def export_evidence(
        cls,
        evidence_list: List[Dict[str, Any]],
        format: str = 'csv',
        filename: Optional[str] = None
    ) -> str:
        """
        导出证据清单

        Args:
            evidence_list: 证据列表
            format: 导出格式
            filename: 文件名

        Returns:
            导出文件路径
        """
        headers = ["类型", "ID", "时间", "相关方", "内容/描述", "命中关键词", "评分", "涉嫌罪名"]
        data = []
        for e in evidence_list:
            e_type = e.get('type', '')
            if e_type == '通讯记录':
                related = f"{e.get('initiator', '')}->{e.get('receiver', '')}"
            else:
                related = f"{e.get('sender', '')}->{e.get('receiver', '')}"

            data.append({
                "类型": e_type,
                "ID": e.get('id', ''),
                "时间": e.get('time', ''),
                "相关方": related,
                "内容/描述": e.get('content', '') or e.get('description', ''),
                "命中关键词": ",".join(e.get('hit_keywords', [])),
                "评分": e.get('score', 0),
                "涉嫌罪名": e.get('crime_type', '')
            })

        if format == 'excel':
            return cls.export_to_excel(data, headers, filename)
        return cls.export_to_csv(data, headers, filename)

    @classmethod
    def export_case_report(
        cls,
        case_data: Dict[str, Any],
        chain_analysis: Dict[str, Any],
        evidence_inventory: Dict[str, Any],
        format: str = 'csv',
        filename: Optional[str] = None
    ) -> str:
        """
        导出案件综合报告（多Sheet）

        Args:
            case_data: 案件数据
            chain_analysis: 上下游分析数据
            evidence_inventory: 证据清单数据
            format: 导出格式
            filename: 文件名

        Returns:
            导出文件路径
        """
        if filename is None:
            case_no = case_data.get('case', {}).get('case_no', 'unknown')
            filename = f"case_report_{case_no}"

        if format == 'excel' and HAS_OPENPYXL:
            return cls._export_case_report_excel(case_data, chain_analysis, evidence_inventory, filename)
        else:
            # 降级为多个CSV
            return cls._export_case_report_csv(case_data, chain_analysis, evidence_inventory, filename)

    @classmethod
    def _export_case_report_csv(cls, case_data, chain_analysis, evidence_inventory, filename) -> str:
        """导出案件报告为CSV（多文件）"""
        base_filename = filename
        output_dir = cls.DEFAULT_EXPORT_DIR
        os.makedirs(output_dir, exist_ok=True)

        # 导出交易记录
        transactions = case_data.get('transactions', [])
        if transactions:
            cls.export_transactions(transactions, 'csv', f"{base_filename}_transactions")

        # 导出证据清单
        evidence_list = evidence_inventory.get('evidence_list', [])
        if evidence_list:
            cls.export_evidence(evidence_list, 'csv', f"{base_filename}_evidence")

        return output_dir

    @classmethod
    def _export_case_report_excel(cls, case_data, chain_analysis, evidence_inventory, filename) -> str:
        """导出案件报告为Excel（多Sheet）"""
        if not HAS_OPENPYXL:
            return cls._export_case_report_csv(case_data, chain_analysis, evidence_inventory, filename)

        filepath = os.path.join(cls.DEFAULT_EXPORT_DIR, f"{filename}.xlsx")
        os.makedirs(cls.DEFAULT_EXPORT_DIR, exist_ok=True)

        wb = openpyxl.Workbook()

        # Sheet 1: 案件概况
        ws_summary = wb.active
        ws_summary.title = "案件概况"
        case = case_data.get('case', {})
        summary_data = [
            ("案件编号", case.get('case_no', '')),
            ("嫌疑人姓名", case.get('suspect_name', '')),
            ("涉案品牌", case.get('brand', '')),
            ("涉案金额", case.get('amount', 0)),
        ]
        for row_idx, (key, value) in enumerate(summary_data, 1):
            ws_summary.cell(row=row_idx, column=1, value=key)
            ws_summary.cell(row=row_idx, column=2, value=value)

        # Sheet 2: 可疑线索
        ws_clues = wb.create_sheet("可疑线索")
        clue_headers = ["ID", "类型", "证据原文", "评分", "涉嫌罪名", "严重程度"]
        ws_clues.append(clue_headers)
        for clue in case_data.get('suspicious_clues', []):
            ws_clues.append([
                clue.get('id', ''),
                clue.get('clue_type', ''),
                clue.get('evidence_text', ''),
                clue.get('score', 0),
                clue.get('crime_type', ''),
                clue.get('severity_level', '')
            ])

        # Sheet 3: 证据清单
        ws_evidence = wb.create_sheet("证据清单")
        evidence_headers = ["类型", "ID", "时间", "相关方", "内容", "命中关键词", "评分"]
        ws_evidence.append(evidence_headers)
        for ev in evidence_inventory.get('evidence_list', []):
            ws_evidence.append([
                ev.get('type', ''),
                ev.get('id', ''),
                ev.get('time', ''),
                f"{ev.get('initiator', '')}->{ev.get('receiver', '')}" if ev.get('type') == '通讯记录' else f"{ev.get('sender', '')}->{ev.get('receiver', '')}",
                ev.get('content', '') or ev.get('description', ''),
                ",".join(ev.get('hit_keywords', [])),
                ev.get('score', 0)
            ])

        wb.save(filepath)
        return filepath
