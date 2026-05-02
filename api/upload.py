"""
数据导入API路由
"""

from fastapi import APIRouter, UploadFile, File, Form, HTTPException
from fastapi.responses import StreamingResponse
from typing import Optional, List, Any
from datetime import datetime
import tempfile
import os
import io
import csv
from urllib.parse import quote

from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment

from services.upload_service import UploadService, TableFormatError
from services.clean_service import CleanService
from services.case_service import CaseService
from services.suspicion_detector import SuspicionDetector
from services.transaction_cross_validator import TransactionCrossValidator
from models.database import Case

router = APIRouter(prefix="/api/upload", tags=["数据导入"])

# 微信 CSV 特征列（与 wechat_parser.WECHAT_SIGNATURE_COLUMNS 保持一致）
_WECHAT_SIGNATURE = {"way", "sender", "senderName", "mediaType", "isDelete"}


# ======================== 自动类型检测 ========================

def _read_headers(temp_path: str, suffix: str) -> List[Any]:
    """读取文件表头行（CSV 或 Excel）"""
    if suffix in (".xlsx", ".xls"):
        wb = load_workbook(temp_path, data_only=True)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        wb.close()
        return headers
    else:
        with open(temp_path, encoding="utf-8-sig") as f:
            reader = csv.reader(f)
            return next(reader, [])


def _detect_upload_type(headers: List[Any]) -> str:
    """
    根据表头自动识别上传数据类型。

    检测顺序：微信格式 → 资金流水 → 通讯记录 → 物流记录
    """
    normalized_headers = UploadService._normalize_headers(headers)
    normalized = set(normalized_headers)

    if not normalized:
        raise TableFormatError("无法识别文件类型：文件缺少表头")

    # 1. 微信格式（特征列完全匹配，必须在通用通讯检测之前）
    raw_set = {str(h).strip() for h in headers if h is not None}
    if _WECHAT_SIGNATURE.issubset(raw_set):
        return "communications"

    def has_any(*candidates: str) -> bool:
        return any(c in normalized for c in candidates)

    # 2. 资金流水
    if (
        has_any("交易发生时间")
        and has_any("打款方", "打款方 (账号/姓名)")
        and has_any("收款方", "收款方 (账号/姓名)")
        and has_any("交易金额", "交易金额 (元)")
    ):
        return "transactions"

    # 3. 通讯记录（标准模板）
    if (
        has_any("联络时间")
        and has_any("发起方 (微信号/姓名)")
        and has_any("接收方 (微信号/姓名)")
        and has_any("聊天内容")
    ):
        return "communications"

    # 4. 物流记录
    if (
        has_any("发货时间")
        and has_any("发件人/网点")
        and has_any("收件人/地址")
    ):
        return "logistics"

    raise TableFormatError(
        "无法识别文件类型，请检查列名是否与模板一致。"
        "支持的类型：资金流水、通讯记录（含微信导出）、物流记录"
    )


# ======================== 保存函数 ========================

def _save_transactions(case, records: list) -> tuple:
    """保存资金流水记录，返回 (saved_count, extra_dict)"""
    from models.database import Transaction

    saved = 0
    for r in records:
        Transaction.create(
            case=case,
            transaction_time=r.get("transaction_time", datetime.now()),
            payer=r.get("payer", ""),
            payee=r.get("payee", ""),
            amount=r.get("amount", 0),
            payment_method=r.get("payment_method"),
            remark=r.get("remark"),
        )
        saved += 1
    return saved, {"case_amount": float(CaseService.recalculate_case_amount(case.id))}


def _save_communications(case, records: list) -> tuple:
    """保存通讯记录，如有微信嵌入转账则一并写入。返回 (saved_count, extra_dict)"""
    from models.database import Communication, Transaction

    saved = 0
    for r in records:
        Communication.create(
            case=case,
            communication_time=r.get("communication_time") or datetime.now(),
            initiator=r.get("initiator", ""),
            receiver=r.get("receiver", ""),
            content=r.get("content"),
            media_type=r.get("media_type"),
            is_deleted=r.get("is_deleted", False),
            raw_content=r.get("raw_content"),
        )
        saved += 1

    # 微信嵌入的转账记录
    wechat_txns = UploadService.get_wechat_transactions()
    txn_saved = 0
    for txn in wechat_txns:
        if txn.get("transaction_time") and txn.get("amount", 0) > 0:
            Transaction.create(
                case=case,
                transaction_time=txn.get("transaction_time", datetime.now()),
                payer=txn.get("payer", ""),
                payee=txn.get("payee", ""),
                amount=txn.get("amount", 0),
                payment_method=txn.get("payment_method"),
                remark=txn.get("remark"),
            )
            txn_saved += 1

    extra = {}
    if txn_saved > 0:
        extra["extracted_transactions"] = txn_saved
        extra["case_amount"] = float(CaseService.recalculate_case_amount(case.id))
    return saved, extra


def _save_logistics(case, records: list) -> tuple:
    """保存物流记录，返回 (saved_count, extra_dict)"""
    from models.database import Logistics

    saved = 0
    for r in records:
        Logistics.create(
            case=case,
            shipping_time=r.get("shipping_time", datetime.now()),
            tracking_no=r.get("tracking_no"),
            sender=r.get("sender", ""),
            sender_address=r.get("sender_address"),
            receiver=r.get("receiver", ""),
            receiver_address=r.get("receiver_address"),
            description=r.get("description"),
            weight=r.get("weight"),
        )
        saved += 1
    return saved, {}


# ======================== 统一上传核心 ========================

async def _handle_upload(
    file: UploadFile,
    case_id: Optional[int] = None,
    case_no: Optional[str] = None,
    forced_type: Optional[str] = None,
):
    """
    统一上传处理：校验 → 查案件 → 写临时文件 → 识别类型 → 解析 → 清洗 → 保存 → 后处理

    Args:
        file: 上传文件
        case_id: 案件ID
        case_no: 案件编号
        forced_type: 强制指定类型（None 时自动识别），用于旧端点兼容
    """

    # 1. 扩展名校验
    if not file.filename or not file.filename.endswith((".xlsx", ".xls", ".csv")):
        raise HTTPException(status_code=400, detail="仅支持 Excel 或 CSV 文件")

    # 2. 查找案件
    case = None
    if case_id:
        case = Case.get_or_none(Case.id == case_id)
    elif case_no:
        case = Case.get_or_none(Case.case_no == case_no)
    if not case:
        raise HTTPException(status_code=404, detail="案件不存在")

    temp_path = None
    try:
        # 3. 写临时文件
        contents = await file.read()
        suffix = os.path.splitext(file.filename)[1] or ".csv"
        with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
            tmp.write(contents)
            temp_path = tmp.name

        # 4. 读取表头 → 自动识别类型
        headers = _read_headers(temp_path, suffix)

        if forced_type:
            data_type = forced_type
            # 微信格式仍然需要检测（用于通讯导入时判断是否需要 wechat 标记）
            is_wechat = (
                forced_type == "communications"
                and suffix in (".csv", ".txt")
                and UploadService._detect_wechat_format(temp_path)
            )
        else:
            data_type = _detect_upload_type(headers)
            is_wechat = (
                data_type == "communications"
                and suffix in (".csv", ".txt")
                and UploadService._detect_wechat_format(temp_path)
            )

        # 5. 按类型解析 → 清洗 → 保存
        if data_type == "transactions":
            records = UploadService.parse_transactions(temp_path, case.id)
            cleaned = CleanService.clean_transactions(records)
            saved_count, extra = _save_transactions(case, cleaned)
            label = "资金流水"
        elif data_type == "communications":
            records = UploadService.parse_communications(temp_path, case.id)
            cleaned = CleanService.clean_communications(records)
            saved_count, extra = _save_communications(case, cleaned)
            label = "通讯记录"
        elif data_type == "logistics":
            records = UploadService.parse_logistics(temp_path, case.id)
            cleaned = CleanService.clean_logistics(records)
            saved_count, extra = _save_logistics(case, cleaned)
            label = "物流记录"
        else:
            raise HTTPException(status_code=400, detail=f"不支持的数据类型: {data_type}")

        # 6. 统一后处理
        inferred_fields = CaseService.auto_update_inferred_fields(case.id)
        person_sync = CaseService.sync_case_persons_to_db(case.id)

        clue_results = SuspicionDetector.detect_all(case.id)
        total_clues = (
            len(clue_results["suspicion_clues"])
            + len(clue_results["price_clues"])
            + len(clue_results["role_clues"])
        )

        cross_results = TransactionCrossValidator.validate(case.id)

        # 7. 构建响应
        result = {
            "success": True,
            "data_type": data_type,
            "message": f"成功导入 {saved_count} 条{label}",
            "case_id": case.id,
            "case_no": case.case_no,
            "case_suspect_name": (
                inferred_fields.get("suspect_name") if inferred_fields else None
            ),
            "case_brand": inferred_fields.get("brand") if inferred_fields else None,
            "person_sync": person_sync,
            "total_records": len(records),
            "saved_records": saved_count,
            "clues_generated": total_clues,
            "cross_anomalies": len(cross_results),
        }

        # transactions 专用：重算金额
        if data_type == "transactions":
            result["case_amount"] = extra.get("case_amount", 0)

        # communications 专用
        if data_type == "communications":
            if is_wechat:
                result["format_detected"] = "wechat"
            if "extracted_transactions" in extra:
                result["extracted_transactions"] = extra["extracted_transactions"]
            if "case_amount" in extra:
                result["case_amount"] = extra["case_amount"]

        return result

    except TableFormatError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"导入失败: {str(e)}")
    finally:
        if temp_path and os.path.exists(temp_path):
            os.unlink(temp_path)


# ======================== 上传路由 ========================


@router.post("/data")
async def upload_data(
    file: UploadFile = File(...),
    case_id: Optional[int] = Form(None),
    case_no: Optional[str] = Form(None),
):
    """
    统一上传端点（推荐）

    根据文件表头自动识别类型：
    - 微信特征列 → 通讯记录（微信格式，含转账提取）
    - 交易发生时间 + 打款方 + 收款方 + 交易金额 → 资金流水
    - 联络时间 + 发起方 + 接收方 + 聊天内容 → 通讯记录
    - 发货时间 + 发件人/网点 + 收件人/地址 → 物流记录
    """
    return await _handle_upload(file, case_id, case_no)


@router.post("/transactions")
async def upload_transactions(
    file: UploadFile = File(...),
    case_id: Optional[int] = Form(None),
    case_no: Optional[str] = Form(None),
):
    """上传资金流水（兼容端点，前端仍在使用）"""
    return await _handle_upload(file, case_id, case_no, forced_type="transactions")


@router.post("/communications")
async def upload_communications(
    file: UploadFile = File(...),
    case_id: Optional[int] = Form(None),
    case_no: Optional[str] = Form(None),
):
    """上传通讯记录（兼容端点，前端仍在使用）"""
    return await _handle_upload(file, case_id, case_no, forced_type="communications")


@router.post("/logistics")
async def upload_logistics(
    file: UploadFile = File(...),
    case_id: Optional[int] = Form(None),
    case_no: Optional[str] = Form(None),
):
    """上传物流记录（兼容端点，前端仍在使用）"""
    return await _handle_upload(file, case_id, case_no, forced_type="logistics")


# ======================== 模板下载 ========================


def _generate_template_xlsx(headers: list, sample_rows: list, sheet_title: str) -> io.BytesIO:
    """生成标准导入模板 XLSX 文件"""
    from openpyxl import Workbook as _Wb

    wb = _Wb()
    ws = wb.active
    ws.title = sheet_title

    header_font = Font(bold=True, size=11)
    header_align = Alignment(horizontal="center", vertical="center")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.alignment = header_align

    for row_idx, row_data in enumerate(sample_rows, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    for col_idx, header in enumerate(headers, 1):
        max_width = len(header) * 2
        for row_data in sample_rows:
            cell_text = str(row_data[col_idx - 1]) if col_idx - 1 < len(row_data) else ""
            max_width = max(max_width, len(cell_text) * 2)
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_width + 4, 50)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


_TEMPLATE_DEFS = {
    "transactions": {
        "filename": "资金流水导入模板.xlsx",
        "sheet_title": "资金流水",
        "headers": [
            "交易发生时间",
            "打款方 (账号/姓名)",
            "收款方 (账号/姓名)",
            "交易金额 (元)",
            "支付方式",
            "交易备注 / 转账留言",
        ],
        "sample_rows": [
            ["2026-01-01 10:00", "张三", "李四", "50000.00", "银行卡", "货款"],
            ["2026-01-02 14:30", "王五", "赵六", "12000.00", "微信", "订金"],
        ],
    },
    "communications": {
        "filename": "通讯记录导入模板.xlsx",
        "sheet_title": "通讯记录",
        "headers": [
            "联络时间",
            "发起方 (微信号/姓名)",
            "接收方 (微信号/姓名)",
            "聊天内容",
        ],
        "sample_rows": [
            ["2026-01-01 10:00", "张三", "李四", "货收到了吗"],
            ["2026-01-01 10:05", "李四", "张三", "收到了，质量不错"],
        ],
    },
    "logistics": {
        "filename": "物流记录导入模板.xlsx",
        "sheet_title": "物流记录",
        "headers": [
            "发货时间",
            "快递单号",
            "发件人/网点",
            "收件人/地址",
            "寄件物品描述",
            "包裹重量(公斤)",
        ],
        "sample_rows": [
            ["2026-01-01 10:00", "SF123456", "张三 (东莞某工业区)", "李四 (杭州余杭某村)", "汽车配件/大灯", "8.0"],
            ["2026-01-02 14:00", "YT789012", "王五 (广州白云)", "赵六 (长沙)", "机油", "45.0"],
        ],
    },
}


@router.get("/template/{type}")
async def download_template(type: str):
    """统一下载导入模板（推荐）"""
    tpl = _TEMPLATE_DEFS.get(type)
    if not tpl:
        raise HTTPException(
            status_code=400,
            detail=f"不支持的模板类型: {type}，可选: transactions / communications / logistics",
        )
    xlsx = _generate_template_xlsx(tpl["headers"], tpl["sample_rows"], tpl["sheet_title"])
    return StreamingResponse(
        xlsx,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(tpl['filename'])}"},
    )


@router.get("/template/transactions")
async def download_transactions_template():
    """下载资金流水导入模板（兼容端点）"""
    return await download_template("transactions")


@router.get("/template/communications")
async def download_communications_template():
    """下载通讯记录导入模板（兼容端点）"""
    return await download_template("communications")


@router.get("/template/logistics")
async def download_logistics_template():
    """下载物流记录导入模板（兼容端点）"""
    return await download_template("logistics")
